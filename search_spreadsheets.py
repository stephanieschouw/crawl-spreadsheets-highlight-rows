# Search excel spreadsheets for text, highlight any rows where they appear

from tkinter import filedialog
from tkinter import *
import os
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re

root = Tk()
root.withdraw()
# Go to the folder you want to crawl
folder_selected = filedialog.askdirectory()
# Go to the folder you want to dump the spreadsheets
folder_dump_loc = filedialog.askdirectory()

# Creates fill pattern for cell
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Gets user input on search term
string_val = input("Enter the text you want to search for: ")

highlight_rows = []

# Iterates through spreadsheets in folder
for filename in os.listdir(folder_selected):
    filepath = os.path.join(folder_selected, filename)
    if os.path.isfile(filepath):
        
        # Checks if search term is found
        found = False

        # Read spreadsheet
        ms = op.load_workbook(filepath)
        ws = ms.active

        # Iterate through all columns and rows
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Looks for search term in the cell value
                if string_val.lower() in str(cell.value).lower():
                    found = True
                    cell_position = str(get_column_letter(cell.column)) + str(cell.row)
                    split_cell_position = re.compile("([a-zA-Z]+)([0-9]+)")
                    res = split_cell_position.match(cell_position).groups()

                    # Add row to list
                    if str(res[1]) not in highlight_rows:
                        highlight_rows.append(str(res[1]))

        # Iterate through all columns and rows
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell_position = str(get_column_letter(cell.column)) + str(cell.row)

                split_cell_position = re.compile("([a-zA-Z]+)([0-9]+)")
                res = split_cell_position.match(cell_position).groups()

                # If row of cell in the list, highlight row
                if str(res[1]) in highlight_rows:
                    ws[cell_position].fill = yellowFill
       
        # If search term is found then create a new spreadsheet in the folder dump location
        if found:
            print(filename + ': text found')
            ms.save(os.path.join(folder_dump_loc, filename))
